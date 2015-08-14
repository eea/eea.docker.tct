import csv
import json
import cStringIO, codecs
from openpyxl import load_workbook
from django.core.management.base import BaseCommand


class UnicodeWriter:
    """
    A CSV writer which will write rows to CSV file "f",
    which is encoded in the given encoding.
    """

    def __init__(self, f, dialect=csv.excel, encoding="utf-8", **kwds):
        # Redirect output to a queue
        self.queue = cStringIO.StringIO()
        self.writer = csv.writer(self.queue, dialect=dialect, **kwds)
        self.stream = f
        self.encoder = codecs.getincrementalencoder(encoding)()

    def writerow(self, row):
        self.writer.writerow([s.encode("utf-8") for s in row])
        # Fetch UTF-8 output from the queue ...
        data = self.queue.getvalue()
        data = data.decode("utf-8")
        # ... and reencode it into the target encoding
        data = self.encoder.encode(data)
        # write to the target stream
        self.stream.write(data)
        # empty queue
        self.queue.truncate(0)

    def writerows(self, rows):
        for row in rows:
            self.writerow(row)


class Command(BaseCommand):
    def handle(self, *args, **options):
        filename = args[0]

        wb = load_workbook(filename=filename)
        first_sheet = wb.get_sheet_names()[0]
        worksheet = wb.get_sheet_by_name(first_sheet)

        data = []
        for row in worksheet.iter_rows():
            data.append([c.value or '' for c in row[1:]])

        header = data[0]
        dict_data = {}
        for row in data[1:]:
            country = row[0]
            dict_data.setdefault(country, [])
            dict_data[country].append(row[1:])

        for country, rows in dict_data.items():
            with open('{}.csv'.format(country), 'w') as fout:
                csv_writer = UnicodeWriter(fout)
                csv_writer.writerow(header)
                csv_writer.writerows(rows)
            print("{}.csv written!".format(country))
